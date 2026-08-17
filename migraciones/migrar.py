"""
Convierte las hojas CARLI, SERGIO y EVENTUALES de la planilla mensual
histórica al formato exacto de la plantilla de importación de SaniFlow.

No escribe nada en la base de datos — genera un .xlsx que después se
sube a mano por la pantalla "Importar clientes desde Excel", donde pasa
por la misma revisión (ok/advertencia/error/duplicado) que cualquier
carga masiva.
"""
import re
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill

ENCABEZADOS = [
    "Nombre", "Tipo (particular/comercio/industria)", "Teléfono", "Email",
    "Dirección", "Localidad", "Notas",
    "Servicio recurrente (dejar vacío si es eventual)",
    "Frecuencia (semanal/quincenal/mensual/trimestral/semestral/anual)",
    "Precio acordado (opcional)",
]

SERVICIO_GENERICO = "Control de plagas"  # ajustar a mano en el Excel si corresponde otro


def _es_empresa(texto: str) -> bool:
    t = (texto or "").upper()
    return any(marca in t for marca in ["S.R.L", "S.A.", "S.A.S", "S.A.I.C", "COOP", "HOSPITAL", "UNIVERSIDAD", "COLEGIO", "ESCUELA", "ESC ", "MUNICIP"])


def _telefono_valido(valor) -> str:
    if valor is None:
        return ""
    s = str(valor).strip()
    solo_digitos = re.sub(r"[^\d]", "", s)
    if 6 <= len(solo_digitos) <= 15:
        return s
    return ""


def extraer_carli_sergio(ws, col_num, col_nombre, col_precio, col_direccion, col_cuit, fuente: str):
    filas = []
    for r in range(4, ws.max_row + 1):
        numero = ws.cell(row=r, column=col_num).value
        nombre = ws.cell(row=r, column=col_nombre).value
        if not isinstance(numero, (int, float)):
            continue  # filtra filas de resumen (Anterior, TOTAL, Diferencia, etc.)
        if not nombre or not str(nombre).strip():
            continue
        nombre = str(nombre).strip()
        if nombre.lower() == "varios":
            continue

        precio = ws.cell(row=r, column=col_precio).value
        direccion = ws.cell(row=r, column=col_direccion).value
        cuit = ws.cell(row=r, column=col_cuit).value

        notas = f"Migrado desde planilla histórica ({fuente}, mayo 2026)."
        if cuit:
            notas += f" CUIT/CUIL: {cuit}."

        filas.append({
            "nombre": nombre,
            "tipo": "comercio",  # clientes con facturación/CUIT, mayormente empresas
            "telefono": "",
            "email": "",
            "direccion": str(direccion).strip() if direccion else "",
            "localidad": "",
            "notas": notas,
            "servicio": SERVICIO_GENERICO,
            "frecuencia": "mensual",
            "precio_acordado": str(precio) if isinstance(precio, (int, float)) and precio > 0 else "",
        })
    return filas


ETIQUETAS_NO_CLIENTE = {"total", "varios", "anterior", "nuevo", "diferencia", "saldo", "subtotal"}


def extraer_eventuales(ws):
    filas = []
    for r in range(4, ws.max_row + 1):
        nombre = ws.cell(row=r, column=5).value  # E
        if not nombre or not str(nombre).strip():
            continue
        nombre = str(nombre).strip()
        if nombre.lower() in ETIQUETAS_NO_CLIENTE:
            continue

        direccion = ws.cell(row=r, column=6).value  # F
        tel_raw = ws.cell(row=r, column=7).value  # G

        filas.append({
            "nombre": nombre,
            "tipo": "comercio" if _es_empresa(nombre) or _es_empresa(direccion) else "particular",
            "telefono": _telefono_valido(tel_raw),
            "email": "",
            "direccion": str(direccion).strip() if direccion else "",
            "localidad": "",
            "notas": "Cliente eventual migrado desde planilla histórica (mayo 2026).",
            "servicio": "",
            "frecuencia": "",
            "precio_acordado": "",
        })
    return filas


def main(archivo_origen: str, archivo_salida: str):
    wb_origen = openpyxl.load_workbook(archivo_origen, data_only=True)

    carli = extraer_carli_sergio(
        wb_origen["CARLI"], col_num=2, col_nombre=3, col_precio=4, col_direccion=16, col_cuit=17, fuente="CARLI"
    )
    sergio = extraer_carli_sergio(
        wb_origen["SERGIO"], col_num=2, col_nombre=3, col_precio=4, col_direccion=15, col_cuit=16, fuente="SERGIO"
    )
    eventuales = extraer_eventuales(wb_origen["EVENTUALES"])

    todas = carli + sergio + eventuales

    # Si el mismo nombre aparece más de una vez (ej: dos contratos o ubicaciones
    # distintas facturados bajo el mismo cliente en la planilla original), el
    # sistema de importación solo se queda con el primero y descarta el resto
    # como "duplicado" — para no perder esa segunda línea de facturación
    # silenciosamente, se renombra con un sufijo y se deja anotado para revisar.
    conteo = {}
    for fila in todas:
        clave = fila["nombre"].strip().lower()
        conteo[clave] = conteo.get(clave, 0) + 1
    vistos = {}
    for fila in todas:
        clave = fila["nombre"].strip().lower()
        if conteo[clave] > 1:
            vistos[clave] = vistos.get(clave, 0) + 1
            if vistos[clave] > 1:
                original = fila["nombre"]
                fila["nombre"] = f"{original} ({vistos[clave]})"
                fila["notas"] = f"REVISAR: puede ser el mismo cliente que \"{original}\", con otro contrato/ubicación en la planilla original. " + fila["notas"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.append(ENCABEZADOS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F5C56")

    for fila in todas:
        ws.append([
            fila["nombre"], fila["tipo"], fila["telefono"], fila["email"],
            fila["direccion"], fila["localidad"], fila["notas"],
            fila["servicio"], fila["frecuencia"], fila["precio_acordado"],
        ])

    anchos = [28, 22, 16, 22, 32, 18, 55, 26, 24, 18]
    for col, ancho in zip("ABCDEFGHIJ", anchos):
        ws.column_dimensions[col].width = ancho

    wb.save(archivo_salida)
    print(f"CARLI: {len(carli)} clientes")
    print(f"SERGIO: {len(sergio)} clientes")
    print(f"EVENTUALES: {len(eventuales)} clientes")
    print(f"TOTAL: {len(todas)} filas escritas en {archivo_salida}")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
