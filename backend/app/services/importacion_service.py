import io
import unicodedata
from dataclasses import dataclass, field
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.cliente import Cliente, TipoCliente
from app.models.servicio import Servicio
from app.models.contrato import Contrato, FrecuenciaContrato
from app.schemas.importacion import FilaImportacion

# Cómo reconocer cada columna de la planilla sin depender de que el
# encabezado esté escrito exactamente igual (el padre del usuario, u otra
# persona no técnica, puede editar la planilla y no poner el texto exacto).
_CAMPOS_CLAVE = {
    "nombre": "nombre",
    "tipo": "tipo",
    "tel": "telefono",
    "mail": "email",
    "direc": "direccion",
    "local": "localidad",
    "nota": "notas",
    "servicio": "servicio",
    "frecuen": "frecuencia",
    "precio": "precio_acordado",
}

_FRECUENCIAS = {
    "semanal": FrecuenciaContrato.semanal,
    "quincenal": FrecuenciaContrato.quincenal,
    "mensual": FrecuenciaContrato.mensual,
    "trimestral": FrecuenciaContrato.trimestral,
    "semestral": FrecuenciaContrato.semestral,
    "anual": FrecuenciaContrato.anual,
    "demanda": FrecuenciaContrato.a_demanda,
}


def _normalizar(s: str) -> str:
    s = (s or "").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return s


def _match_tipo(valor: str) -> tuple[TipoCliente, str | None]:
    v = _normalizar(valor)
    if "comerc" in v:
        return TipoCliente.comercio, None
    if "indus" in v:
        return TipoCliente.industria, None
    if not v or "particular" in v:
        return TipoCliente.particular, None
    return TipoCliente.particular, f'Tipo "{valor}" no reconocido, se cargó como Particular.'


def _match_frecuencia(valor: str) -> tuple[FrecuenciaContrato | None, str | None]:
    v = _normalizar(valor)
    if not v:
        return None, None
    for clave, enum_val in _FRECUENCIAS.items():
        if clave in v:
            return enum_val, None
    return None, f'Frecuencia "{valor}" no reconocida.'


@dataclass
class FilaProcesada:
    fila: int
    raw: FilaImportacion
    estado: str  # "ok" | "advertencia" | "error" | "duplicado"
    mensajes: list[str] = field(default_factory=list)
    nombre: str = ""
    tipo: TipoCliente = TipoCliente.particular
    telefono: str | None = None
    email: str | None = None
    direccion: str | None = None
    localidad: str | None = None
    notas: str | None = None
    servicio_id: int | None = None
    servicio_nombre: str | None = None
    frecuencia: FrecuenciaContrato | None = None
    precio_acordado: float | None = None


def parsear_excel(contenido: bytes) -> list[FilaImportacion]:
    """Lee el archivo subido y devuelve una fila por cada renglón con datos,
    ignorando filas totalmente vacías (para no romperse si alguien dejó
    espacio de más en la planilla)."""
    wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    ws = wb.worksheets[0]
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return []

    columnas: dict[int, str] = {}
    for idx, valor in enumerate(filas[0]):
        if not valor:
            continue
        norm = _normalizar(str(valor))
        for clave, campo in _CAMPOS_CLAVE.items():
            if clave in norm and campo not in columnas.values():
                columnas[idx] = campo
                break

    resultado = []
    for fila in filas[1:]:
        if fila is None or all(v is None or str(v).strip() == "" for v in fila):
            continue
        registro = {campo: "" for campo in _CAMPOS_CLAVE.values()}
        for idx, campo in columnas.items():
            if idx < len(fila) and fila[idx] is not None:
                registro[campo] = str(fila[idx]).strip()
        resultado.append(FilaImportacion(**registro))
    return resultado


def procesar_fila(db: Session, numero_fila: int, raw: FilaImportacion) -> FilaProcesada:
    """Valida e interpreta una fila. No escribe nada en la base — eso lo
    hace `confirmar`, que llama a esta misma función para no duplicar
    la lógica de validación entre la vista previa y la carga real."""
    nombre = raw.nombre.strip()
    if not nombre:
        return FilaProcesada(
            fila=numero_fila, raw=raw, estado="error",
            mensajes=["Falta el nombre del cliente — esta fila no se puede cargar."],
        )

    existente = db.query(Cliente).filter(func.lower(Cliente.nombre) == nombre.lower()).first()
    if existente:
        return FilaProcesada(
            fila=numero_fila, raw=raw, nombre=nombre, estado="duplicado",
            mensajes=[f'Ya existe un cliente llamado "{nombre}" — esta fila no se vuelve a cargar.'],
        )

    mensajes: list[str] = []
    tipo, msg_tipo = _match_tipo(raw.tipo)
    if msg_tipo:
        mensajes.append(msg_tipo)

    servicio_id = None
    servicio_nombre = None
    frecuencia = None
    precio_acordado = None
    servicio_txt = raw.servicio.strip()

    if servicio_txt:
        servicio = (
            db.query(Servicio)
            .filter(func.lower(Servicio.nombre).like(f"%{servicio_txt.lower()}%"))
            .first()
        )
        if not servicio:
            mensajes.append(f'Servicio "{servicio_txt}" no encontrado — el cliente se carga como Eventual (sin contrato).')
        else:
            frecuencia, msg_frec = _match_frecuencia(raw.frecuencia)
            if not frecuencia:
                mensajes.append(
                    (msg_frec or "Falta la frecuencia") + " — el cliente se carga como Eventual (sin contrato)."
                )
            else:
                servicio_id = servicio.id
                servicio_nombre = servicio.nombre
                precio_txt = raw.precio_acordado.strip()
                if precio_txt:
                    try:
                        precio_acordado = float(precio_txt.replace(",", "."))
                    except ValueError:
                        mensajes.append(f'Precio "{precio_txt}" inválido, se usa el precio base del servicio.')

    return FilaProcesada(
        fila=numero_fila, raw=raw, estado="advertencia" if mensajes else "ok", mensajes=mensajes,
        nombre=nombre, tipo=tipo,
        telefono=raw.telefono.strip() or None, email=raw.email.strip() or None,
        direccion=raw.direccion.strip() or None, localidad=raw.localidad.strip() or None,
        notas=raw.notas.strip() or None,
        servicio_id=servicio_id, servicio_nombre=servicio_nombre,
        frecuencia=frecuencia, precio_acordado=precio_acordado,
    )


def previsualizar(db: Session, contenido: bytes) -> list[FilaProcesada]:
    filas = parsear_excel(contenido)
    # +2: la fila 1 del Excel es el encabezado, los datos arrancan en la fila 2 —
    # así el número de fila que se muestra coincide con lo que la persona ve al abrir el archivo.
    return [procesar_fila(db, i + 2, raw) for i, raw in enumerate(filas)]


def confirmar(db: Session, filas: list[FilaImportacion]) -> tuple[int, int, int, list[FilaProcesada]]:
    creados = 0
    contratos_creados = 0
    omitidos = 0
    detalles: list[FilaProcesada] = []

    for i, raw in enumerate(filas):
        procesada = procesar_fila(db, i + 2, raw)
        if procesada.estado in ("error", "duplicado"):
            omitidos += 1
            detalles.append(procesada)
            continue

        cliente = Cliente(
            nombre=procesada.nombre, tipo=procesada.tipo, telefono=procesada.telefono,
            email=procesada.email, direccion=procesada.direccion, localidad=procesada.localidad,
            notas=procesada.notas,
        )
        db.add(cliente)
        db.flush()  # necesario para que el chequeo de duplicados de la próxima fila vea a este cliente

        if procesada.servicio_id:
            servicio = db.query(Servicio).filter(Servicio.id == procesada.servicio_id).first()
            contrato = Contrato(
                cliente_id=cliente.id, servicio_id=procesada.servicio_id,
                frecuencia=procesada.frecuencia,
                precio_acordado=procesada.precio_acordado or (servicio.precio_base if servicio else None),
                fecha_inicio=date.today(),
            )
            db.add(contrato)
            db.flush()
            contratos_creados += 1

        creados += 1
        detalles.append(procesada)

    db.commit()
    return creados, contratos_creados, omitidos, detalles


def generar_plantilla() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    encabezados = [
        "Nombre", "Tipo (particular/comercio/industria)", "Teléfono", "Email",
        "Dirección", "Localidad", "Notas",
        "Servicio recurrente (dejar vacío si es eventual)",
        "Frecuencia (semanal/quincenal/mensual/trimestral/semestral/anual)",
        "Precio acordado (opcional)",
    ]
    ws.append(encabezados)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F5C56")

    ws.append([
        "Farmacia del Sol", "comercio", "260 4123456", "", "San Martín 450, Ciudad",
        "San Rafael", "Cliente de ejemplo — se puede borrar esta fila",
        "Desinsectación general", "mensual", "",
    ])
    ws.append([
        "Sra. Beatriz Coria", "particular", "260 4654321", "", "Chile 120",
        "Villa 25 de Mayo", "Cliente eventual de ejemplo: sin servicio ni frecuencia",
        "", "", "",
    ])

    dv_tipo = DataValidation(type="list", formula1='"particular,comercio,industria"', allow_blank=True)
    ws.add_data_validation(dv_tipo)
    dv_tipo.add("B2:B500")

    dv_frec = DataValidation(
        type="list",
        formula1='"semanal,quincenal,mensual,trimestral,semestral,anual"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_frec)
    dv_frec.add("I2:I500")

    anchos = [24, 30, 16, 22, 28, 18, 34, 34, 40, 20]
    for col, ancho in zip("ABCDEFGHIJ", anchos):
        ws.column_dimensions[col].width = ancho

    ayuda = wb.create_sheet("Instrucciones")
    ayuda["A1"] = "Cómo completar esta planilla"
    ayuda["A1"].font = Font(bold=True, size=14)
    lineas = [
        "",
        "1. Completá una fila por cliente, en la hoja \"Clientes\".",
        "2. Solo el Nombre es obligatorio. El resto de las columnas se pueden dejar vacías.",
        "3. Si el cliente NO tiene un servicio periódico (por ejemplo, llamó una sola vez), "
        "dejá vacías las columnas \"Servicio recurrente\" y \"Frecuencia\". Va a quedar cargado "
        "como cliente Eventual.",
        "4. Si el cliente SÍ tiene un servicio periódico (por ejemplo, fumigación todos los "
        "meses), completá \"Servicio recurrente\" con el nombre del servicio y \"Frecuencia\" "
        "con cada cuánto se repite. Va a quedar cargado como cliente Fijo, con su contrato ya armado.",
        "5. Las dos primeras filas son ejemplos — se pueden borrar antes de completar la tuya, "
        "o dejarlas y borrarlas después de revisar la importación.",
        "6. Guardá el archivo y subilo en la pantalla \"Importar clientes desde Excel\". "
        "El sistema muestra una vista previa antes de cargar nada, para poder revisar y corregir "
        "cualquier error antes de confirmar.",
    ]
    for i, linea in enumerate(lineas, start=2):
        ayuda[f"A{i}"] = linea
    ayuda.column_dimensions["A"].width = 110

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
