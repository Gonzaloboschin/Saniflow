import io

import openpyxl


def _armar_excel(filas: list[list[str]]) -> bytes:
    """Arma un .xlsx real en memoria, con el mismo encabezado que usa la plantilla."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "Nombre", "Tipo (particular/comercio/industria)", "Teléfono", "Email",
        "Dirección", "Localidad", "Notas",
        "Servicio recurrente (dejar vacío si es eventual)",
        "Frecuencia (semanal/quincenal/mensual/trimestral/semestral/anual)",
        "Precio acordado (opcional)",
    ])
    for fila in filas:
        ws.append(fila)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _crear_servicio(client, nombre="Desinsectación general", precio=15000):
    r = client.post("/servicios", json={"nombre": nombre, "precio_base": precio})
    assert r.status_code == 201
    return r.json()


def test_plantilla_descargable(client):
    r = client.get("/clientes/importar/plantilla")
    assert r.status_code == 200
    assert "spreadsheet" in r.headers["content-type"]
    # Confirma que el archivo generado es un Excel válido y legible, no basura binaria.
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert "Clientes" in wb.sheetnames
    assert "Instrucciones" in wb.sheetnames


def test_previsualizar_distingue_ok_error_y_duplicado(client):
    _crear_servicio(client)
    client.post("/clientes", json={"nombre": "Cliente Ya Existente"})

    contenido = _armar_excel([
        ["Cliente Nuevo Eventual", "particular", "260123", "", "", "", "", "", "", ""],
        ["", "particular", "", "", "", "", "", "", "", ""],  # sin nombre -> error
        ["Cliente Ya Existente", "particular", "", "", "", "", "", "", "", ""],  # duplicado
        ["Cliente Fijo Mensual", "comercio", "", "", "", "", "", "Desinsectación general", "mensual", ""],
    ])

    r = client.post(
        "/clientes/importar/previsualizar",
        files={"archivo": ("clientes.xlsx", contenido, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    filas = r.json()
    assert len(filas) == 4

    por_nombre = {f["datos"]["nombre"]: f for f in filas}
    assert por_nombre["Cliente Nuevo Eventual"]["estado"] == "ok"
    assert por_nombre["Cliente Nuevo Eventual"]["es_fijo"] is False
    assert por_nombre[""]["estado"] == "error"
    assert por_nombre["Cliente Ya Existente"]["estado"] == "duplicado"
    assert por_nombre["Cliente Fijo Mensual"]["estado"] == "ok"
    assert por_nombre["Cliente Fijo Mensual"]["es_fijo"] is True
    assert por_nombre["Cliente Fijo Mensual"]["servicio_resuelto"] == "Desinsectación general"

    # Ninguna fila debe haber tocado la base todavía — previsualizar es de solo lectura.
    r = client.get("/clientes")
    nombres = {c["nombre"] for c in r.json()}
    assert "Cliente Nuevo Eventual" not in nombres
    assert "Cliente Fijo Mensual" not in nombres


def test_previsualizar_avisa_servicio_inexistente_sin_bloquear(client):
    contenido = _armar_excel([
        ["Cliente Con Servicio Inventado", "particular", "", "", "", "", "", "Servicio Que No Existe", "mensual", ""],
    ])
    r = client.post(
        "/clientes/importar/previsualizar",
        files={"archivo": ("clientes.xlsx", contenido, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    fila = r.json()[0]
    assert fila["estado"] == "advertencia"  # no bloquea, pero avisa
    assert fila["es_fijo"] is False
    assert "no encontrado" in fila["mensajes"][0]


def test_confirmar_crea_clientes_y_contratos_y_omite_invalidos(client):
    _crear_servicio(client, "Fumigación de jardines", 20000)
    client.post("/clientes", json={"nombre": "Cliente Repetido"})

    filas = [
        {"nombre": "Nuevo Eventual", "tipo": "particular"},
        {"nombre": "", "tipo": "particular"},  # se omite: sin nombre
        {"nombre": "Cliente Repetido", "tipo": "particular"},  # se omite: duplicado
        {"nombre": "Nuevo Fijo", "tipo": "comercio", "servicio": "Fumigación de jardines", "frecuencia": "trimestral"},
    ]
    # Completar los campos que el schema espera con default "" para no mandar None
    filas_completas = [
        {**{"nombre": "", "tipo": "", "telefono": "", "email": "", "direccion": "", "localidad": "",
            "notas": "", "servicio": "", "frecuencia": "", "precio_acordado": ""}, **f}
        for f in filas
    ]

    r = client.post("/clientes/importar/confirmar", json=filas_completas)
    assert r.status_code == 200
    resumen = r.json()

    assert resumen["creados"] == 2
    assert resumen["contratos_creados"] == 1
    assert resumen["omitidos"] == 2

    clientes = {c["nombre"]: c for c in client.get("/clientes").json()}
    assert "Nuevo Eventual" in clientes
    assert "Nuevo Fijo" in clientes

    nuevo_fijo_id = clientes["Nuevo Fijo"]["id"]
    contratos = client.get(f"/contratos/cliente/{nuevo_fijo_id}").json()
    assert len(contratos) == 1
    assert contratos[0]["frecuencia"] == "trimestral"
    assert float(contratos[0]["precio_acordado"]) == 20000.0  # usó el precio base del servicio


def test_confirmar_no_duplica_si_dos_filas_del_mismo_archivo_repiten_nombre(client):
    filas = [
        {"nombre": "Cliente Repetido En El Mismo Archivo", "tipo": "particular", "telefono": "", "email": "",
         "direccion": "", "localidad": "", "notas": "", "servicio": "", "frecuencia": "", "precio_acordado": ""},
        {"nombre": "Cliente Repetido En El Mismo Archivo", "tipo": "particular", "telefono": "", "email": "",
         "direccion": "", "localidad": "", "notas": "", "servicio": "", "frecuencia": "", "precio_acordado": ""},
    ]
    r = client.post("/clientes/importar/confirmar", json=filas)
    resumen = r.json()
    assert resumen["creados"] == 1
    assert resumen["omitidos"] == 1
